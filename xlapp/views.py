import requests
from django.shortcuts import render
import openpyxl
import xlsxwriter
# Create your views here.


def index(request):
    if "GET" == request.method:
        return render(request, 'xlapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                api_key = "YOURAPIKEY"
                address = str(cell.value)
                api_response = requests.get('https://maps.googleapis.com/maps/api/geocode/json?address={0}&key={1}'.format(address, api_key))
                api_response_dict = api_response.json()
                if api_response_dict:
                    latitude = api_response_dict['results'][0]['geometry']['location']['lat']
                    longitude = api_response_dict['results'][0]['geometry']['location']['lng']
                    row_data.append(address)
                    row_data.append(latitude)
                    row_data.append(longitude)
            excel_data.append(row_data)
        workbook = xlsxwriter.Workbook('newexcel.xlsx')
        worksheet = workbook.add_worksheet("My sheet")
        scores = excel_data
        row = 0
        col = 0
        for name, score,score1 in (scores):
            worksheet.write(row, col, name)
            worksheet.write(row, col + 1, score)
            worksheet.write(row, col + 2, score1)
            row += 1
        workbook.close()
        return render(request, 'xlapp/index.html', {})